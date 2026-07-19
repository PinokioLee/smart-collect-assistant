"""Template Design Agent(양식 자동 생성) 회귀 테스트.

검증 관점:
  1) 자연어 → 컬럼 스키마 휴리스틱 설계(재현 가능)
  2) 스펙 → 검증 규칙 변환(라운드트립의 단일 출처)
  3) openpyxl 양식 생성(헤더/드롭다운/안내시트)
  4) '생성한 양식 = 검증 계약' 라운드트립 — 같은 규칙으로 회신을 검증
"""

from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from smart_collect.pipeline import run_collection
from smart_collect.tools.template_tools import (
    build_and_save_template,
    design_template_from_intent,
    load_template_spec,
    template_excel_path,
    template_spec_to_validation_rule,
    write_template_excel,
)

PROJECT_INTENT = (
    "프로젝트별 월 실적을 걷을 건데 "
    "프로젝트번호, 담당자, 매출액, 진행상태(정상/지연/보류), 마감일자 받고 싶어"
)


def test_heuristic_design_preserves_order_and_types():
    spec = design_template_from_intent(PROJECT_INTENT, prefer_llm=False)
    assert spec.source == "heuristic"
    names = [c.name for c in spec.columns]
    assert names == ["프로젝트번호", "담당자", "매출액", "진행상태", "마감일자"]

    by = {c.name: c for c in spec.columns}
    assert by["마감일자"].dtype == "date"
    assert by["매출액"].dtype == "number"
    assert by["진행상태"].dtype == "code"
    assert by["진행상태"].allowed_values == ["정상", "지연", "보류"]
    # 번호/담당자 같은 키 컬럼은 필수로 설계
    assert by["프로젝트번호"].required is True
    assert by["담당자"].required is True


def test_code_column_name_not_clipped():
    # '긴급도' 의 끝글자 '도' 가 조사로 오인되어 잘리면 안 된다.
    spec = design_template_from_intent(
        "부서명, 담당자, 요청시스템, 긴급도(상/중/하), 요청일자", prefer_llm=False
    )
    by = {c.name: c for c in spec.columns}
    assert "긴급도" in by
    assert by["긴급도"].dtype == "code"
    assert by["긴급도"].allowed_values == ["상", "중", "하"]


def test_spec_to_validation_rule():
    spec = design_template_from_intent(PROJECT_INTENT, prefer_llm=False)
    rule = template_spec_to_validation_rule(spec)
    assert rule.required_columns == ["프로젝트번호", "담당자"]
    assert rule.date_columns == ["마감일자"]
    assert rule.code_rules == {"진행상태": ["정상", "지연", "보류"]}


def test_write_template_excel_headers_and_dropdown(tmp_path):
    spec = design_template_from_intent(PROJECT_INTENT, prefer_llm=False)
    path = write_template_excel(spec, tmp_path / "양식.xlsx")
    assert path.exists()

    wb = load_workbook(path)
    ws = wb["취합양식"]
    # 헤더가 컬럼명과 '정확히' 일치해야 회신 검증과 어긋나지 않는다.
    headers = [ws.cell(row=1, column=i + 1).value for i in range(len(spec.columns))]
    assert headers == ["프로젝트번호", "담당자", "매출액", "진행상태", "마감일자"]
    # 작성안내 시트 존재
    assert "작성안내" in wb.sheetnames
    # 코드값 컬럼(진행상태)에 드롭다운(DataValidation)이 걸려 있어야 한다.
    dvs = list(ws.data_validations.dataValidation)
    assert dvs, "코드값 컬럼에 데이터 검증(드롭다운)이 없습니다."
    assert any("정상" in (dv.formula1 or "") for dv in dvs)


def test_build_and_save_persists_and_downloadable():
    spec = design_template_from_intent(PROJECT_INTENT, prefer_llm=False)
    info = build_and_save_template(spec)
    tid = info["template_id"]
    # 스펙 재로딩 → 라운드트립 단일 출처 보존
    reloaded = load_template_spec(tid)
    assert reloaded is not None
    assert [c.name for c in reloaded.columns] == [c.name for c in spec.columns]
    # 생성된 엑셀 파일이 다운로드 가능 위치에 존재
    assert template_excel_path(tid) is not None
    assert info["validation_rule"]["code_rules"] == {"진행상태": ["정상", "지연", "보류"]}


def test_roundtrip_generated_template_is_validation_contract(tmp_path):
    """생성한 양식으로 만든 검증 규칙이 그대로 회신 검증에 쓰인다."""
    spec = design_template_from_intent(PROJECT_INTENT, prefer_llm=False)
    rule = template_spec_to_validation_rule(spec)

    # 양식대로 작성했지만 오류 2건이 섞인 회신 파일:
    #  - 2행: 진행상태 '진행중'(허용값 아님) + 담당자 공란(필수 누락)
    #  - 3행: 정상 행
    submit = tmp_path / "submit.xlsx"
    pd.DataFrame(
        [
            {"프로젝트번호": "PRJ-001", "담당자": "", "매출액": "1000000",
             "진행상태": "진행중", "마감일자": "2026-06-30"},
            {"프로젝트번호": "PRJ-002", "담당자": "김담당", "매출액": "2000000",
             "진행상태": "정상", "마감일자": "2026-06-30"},
        ]
    ).to_excel(submit, index=False, engine="openpyxl")

    state = run_collection(
        "TPL-ROUNDTRIP-TEST", "월 실적 취합", "본문",
        [str(submit)], prefer_llm=False,
        rule_override=rule, template_id="TPL-XYZ",
    )

    # 라운드트립 플래그: 생성 양식이 검증 계약으로 고정됨
    assert state.template_locked is True
    assert state.template_id == "TPL-XYZ"
    # 규칙을 재도출하지 않고 양식 규칙을 그대로 사용
    assert state.validation_rules.code_rules == {"진행상태": ["정상", "지연", "보류"]}
    # 허용되지 않은 코드값과 필수값 누락을 잡아냈다
    types = set(state.validation_result.error_types)
    assert "허용되지 않은 코드값" in types
    assert "필수값 누락" in types
    assert state.validation_result.valid_rows == 1
