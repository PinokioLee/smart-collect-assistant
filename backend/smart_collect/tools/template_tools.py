"""Template Design Agent 도구 — '내가 원하는 취합 양식'을 AI가 설계·생성한다.

흐름 (라운드트립):
  자연어 취합 의도
    → design_template_from_intent()  🧠 LLM(또는 ⚙️ 휴리스틱 폴백): 컬럼 스키마 설계
    → write_template_excel()         ⚙️ 결정론: 배포용 엑셀 양식 생성(드롭다운/서식/안내시트)
    → template_spec_to_validation_rule() ⚙️ 결정론: 이 양식이 곧 회신 검증 규칙

핵심 설계 경계는 기존과 동일하다 — LLM 은 '자연어 → 구조' 설계만 하고,
파일 생성과 검증 규칙 파생은 재현 가능한 결정론 코드가 담당한다.
LLM 이 설계한 스펙은 사용자가 검토·수정한 뒤에만 확정된다(Human-in-the-loop).
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..config import TEMPLATE_DIR
from ..state import ColumnSpec, TemplateSpec, ValidationRule

# ---------- 휴리스틱 설계 (LLM 없이 재현 가능) ----------

_DATE_HINTS = ("일자", "날짜", "일시", "기간", "마감", "date")
_NUMBER_HINTS = ("금액", "매출", "원가", "비용", "수량", "개수", "단가", "예산", "인원", "매출액")
_KEY_HINTS = ("번호", "담당자", "부서", "이름", "성명", "프로젝트명", "시스템", "품목", "코드")
# 프로즈(문장) 토큰을 컬럼으로 오인하지 않도록 거르는 동사/서술 힌트
_PROSE_HINTS = ("걷", "받", "싶", "주세", "바랍", "합니다", "입니다", "하고", "하려", "정리", "취합할")
# 컬럼명에서 떼어낼 접미 조사/군더더기
# 주의: '도/가/이/의' 등은 실제 컬럼명(긴급도·단가·차이) 끝글자와 겹치므로 제외한다.
_TRAIL = ("를", "을", "는", "은", "등", "항목", "정보")
_CODE_PAREN = re.compile(r"^([가-힣A-Za-z0-9_ ]+?)\s*[\(（]\s*([^)）]+?)\s*[\)）]\s*$")
# 리스트 앞의 도입 어구(…할 건데)와 뒤의 서술(…받고 싶어)을 잘라 항목 구간만 남긴다.
_LEADIN = re.compile(r"^.*?(?:건데|는데|한데|하려고|하는데|같은|처럼|으로|로서|[:：\-—>.。!?！？])\s*")
_TRAILVERB = re.compile(
    r"\s*(?:받고\s*싶|받고싶|주세요|바랍|취합|정리|부탁|필요|해\s*줘|해줘|만들|알려).*$"
)


def _list_region(text: str) -> str:
    """자유 문장에서 '항목 나열' 구간만 추출한다(도입어구·말미서술 제거)."""
    t = text.strip()
    # 콤마가 있으면 리스트로 보고 도입/말미를 잘라낸다.
    if "," in t or "，" in t:
        t = t.replace("，", ",")
        t = _LEADIN.sub("", t, count=1)
        t = _TRAILVERB.sub("", t)
    return t.strip()


def _clean_name(raw: str) -> str:
    name = (raw or "").strip().strip("·-•*").strip()
    name = re.sub(r"^(그리고|또|및|각|해당)\s+", "", name)
    # 프로즈 힌트를 가진 단어(걷을/받고 등) 제거 후 재조립
    if " " in name and any(h in name for h in _PROSE_HINTS):
        words = [w for w in name.split() if not any(h in w for h in _PROSE_HINTS)]
        name = " ".join(words).strip()
    # 접미 조사/군더더기 반복 제거
    changed = True
    while changed and name:
        changed = False
        for t in _TRAIL:
            if name.endswith(t) and len(name) > len(t):
                name = name[: -len(t)].strip()
                changed = True
    return name.strip()


def _looks_like_field(name: str) -> bool:
    if not name or len(name) > 20:
        return False
    return not any(h in name for h in _PROSE_HINTS)


def _infer_dtype(name: str) -> str:
    low = name.lower()
    if any(h in low for h in _DATE_HINTS):
        return "date"
    if any(h in name for h in _NUMBER_HINTS) or name.endswith(("액", "율", "률", "수")):
        return "number"
    return "text"


def _is_key(name: str) -> bool:
    return any(h in name for h in _KEY_HINTS)


def _example_for(name: str, dtype: str, allowed: list[str] | None = None) -> str:
    if allowed:
        return allowed[0]
    if dtype == "date":
        return "2026-06-30"
    if dtype == "number":
        return "1000000"
    if "번호" in name:
        return "PRJ-001"
    if "담당자" in name or "이름" in name or "성명" in name:
        return "홍길동"
    if "부서" in name:
        return "정보시스템팀"
    return ""


def _default_columns() -> list[ColumnSpec]:
    return [
        ColumnSpec(name="부서명", dtype="text", required=True, example="정보시스템팀"),
        ColumnSpec(name="담당자", dtype="text", required=True, example="홍길동"),
        ColumnSpec(name="항목", dtype="text", required=True, example="항목 내용"),
        ColumnSpec(
            name="긴급도", dtype="code", required=False,
            allowed_values=["상", "중", "하"], example="중",
        ),
        ColumnSpec(name="요청일자", dtype="date", required=False, example="2026-06-30"),
    ]


def design_template_heuristic(intent: str) -> TemplateSpec:
    """LLM 없이 자연어 의도에서 양식 컬럼을 규칙 기반으로 추출한다(재현 가능).

    입력 순서를 보존하며, '이름(값1/값2/값3)' 형태는 코드값 컬럼(드롭다운)으로 인식한다.
    콤마로 항목을 구분한 입력에서 가장 정확하게 동작한다.
    """
    text = (intent or "").replace("\n", " ").strip()
    region = _list_region(text)
    columns: list[ColumnSpec] = []
    seen: set[str] = set()

    # 콤마가 없으면 공백/구분자로도 시도
    tokens = re.split(r"[,\n·、]", region) if ("," in region) else re.split(r"[,\n·、]", text)
    for raw in tokens:
        tok = raw.strip()
        if not tok:
            continue
        m = _CODE_PAREN.match(tok)
        if m:
            name = _clean_name(m.group(1))
            opts = [o.strip() for o in re.split(r"[/·,|、]", m.group(2)) if o.strip()]
            if name and len(opts) >= 2 and name not in seen and _looks_like_field(name):
                columns.append(
                    ColumnSpec(
                        name=name, dtype="code", required=_is_key(name),
                        allowed_values=opts, example=opts[0],
                        description=f"{'/'.join(opts)} 중 하나",
                    )
                )
                seen.add(name)
            continue
        name = _clean_name(tok)
        if not name or name in seen or not _looks_like_field(name):
            continue
        dtype = _infer_dtype(name)
        columns.append(
            ColumnSpec(
                name=name, dtype=dtype, required=_is_key(name),
                example=_example_for(name, dtype),
            )
        )
        seen.add(name)

    if not columns:
        columns = _default_columns()

    dup_keys = [c.name for c in columns if c.required][:3]
    title = "취합 양식"
    return TemplateSpec(
        title=title,
        purpose=text[:60] or None,
        columns=columns,
        duplicate_keys=dup_keys,
        notes=["필수 항목은 반드시 입력해 주세요.", "코드값 컬럼은 드롭다운에서 선택해 주세요."],
        source="heuristic",
    )


def _spec_from_llm_dict(data: dict) -> Optional[TemplateSpec]:
    """LLM JSON dict → TemplateSpec (검증·정규화)."""
    raw_cols = data.get("columns")
    if not isinstance(raw_cols, list) or not raw_cols:
        return None
    columns: list[ColumnSpec] = []
    for c in raw_cols:
        if not isinstance(c, dict) or not str(c.get("name") or "").strip():
            continue
        dtype = str(c.get("dtype") or "text").strip().lower()
        if dtype not in {"text", "date", "number", "code"}:
            dtype = "text"
        allowed = [str(v).strip() for v in (c.get("allowed_values") or []) if str(v).strip()]
        name = str(c["name"]).strip()
        columns.append(
            ColumnSpec(
                name=name,
                dtype=dtype,
                required=bool(c.get("required", False)),
                allowed_values=allowed if dtype == "code" else [],
                date_format=str(c.get("date_format") or "YYYY-MM-DD"),
                example=(str(c.get("example")).strip() if c.get("example") else None)
                or _example_for(name, dtype, allowed) or None,
                description=str(c.get("description")).strip() if c.get("description") else None,
            )
        )
    if not columns:
        return None
    dup = [str(k).strip() for k in (data.get("duplicate_keys") or []) if str(k).strip()]
    if not dup:
        dup = [c.name for c in columns if c.required][:3]
    notes = [str(n).strip() for n in (data.get("notes") or []) if str(n).strip()]
    return TemplateSpec(
        title=str(data.get("title") or "취합 양식").strip(),
        purpose=str(data.get("purpose")).strip() if data.get("purpose") else None,
        deadline=str(data.get("deadline")).strip() if data.get("deadline") else None,
        columns=columns,
        duplicate_keys=[k for k in dup if k in {c.name for c in columns}],
        notes=notes,
        source="llm",
    )


def design_template_from_intent(
    intent: str, *, prefer_llm: bool = True
) -> TemplateSpec:
    """자연어 취합 의도 → 양식 스펙. Azure 키가 있으면 LLM, 아니면 휴리스틱.

    LLM 경로 실패 시 자동으로 휴리스틱으로 폴백한다(무비용·재현 보장).
    """
    if prefer_llm:
        try:
            from ..llm import design_template_with_llm  # 지연 임포트

            data = design_template_with_llm(intent)
            if data is not None:
                spec = _spec_from_llm_dict(data)
                if spec is not None:
                    return spec
        except Exception:  # noqa: BLE001 - 폴백 보장
            pass
    return design_template_heuristic(intent)


# ---------- 스펙 → 검증 규칙 (라운드트립의 핵심) ----------

def template_spec_to_validation_rule(spec: TemplateSpec) -> ValidationRule:
    """생성한 양식 스펙을 그대로 회신 검증 규칙으로 변환한다.

    '보낸 양식 = 검증 계약' — 요청과 검증 기준이 어긋날 수 없도록 단일 출처에서 파생한다.
    """
    return ValidationRule(
        required_columns=[c.name for c in spec.columns if c.required],
        date_columns=[c.name for c in spec.columns if c.dtype == "date"],
        number_columns=[c.name for c in spec.columns if c.dtype == "number"],
        code_rules={
            c.name: list(c.allowed_values)
            for c in spec.columns
            if c.dtype == "code" and c.allowed_values
        },
        duplicate_keys=list(spec.duplicate_keys),
    )


# ---------- 엑셀 양식 생성 (openpyxl, 결정론) ----------

_HEADER_FILL = PatternFill("solid", fgColor="1F3B57")       # 네이비
_REQUIRED_FILL = PatternFill("solid", fgColor="2E7D6B")     # 민트(필수 표시)
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_THIN = Side(style="thin", color="D0D5DD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def write_template_excel(spec: TemplateSpec, path: str | Path) -> Path:
    """양식 스펙으로 배포용 엑셀 파일을 생성한다.

    - 1행 헤더(필수 컬럼은 색으로 구분), 헤더 고정
    - 코드값 컬럼은 드롭다운(DataValidation)으로 잘못된 값 입력을 원천 차단
    - 날짜 컬럼은 날짜 표시 형식 지정
    - 빈 데이터 영역 + '작성안내' 시트(컬럼별 형식/허용값/예시/설명)
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "취합양식"

    cols = spec.columns or _default_columns()

    # 헤더
    for idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=idx, value=col.name)
        cell.font = _HEADER_FONT
        cell.fill = _REQUIRED_FILL if col.required else _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = max(len(col.name) + 4, 14)

    # 데이터 시트에는 예시값을 넣지 않는다. 배포한 양식 여러 개를 취합할 때
    # 예시행이 실제 제출 데이터로 오인·중복되는 사고를 방지한다.
    for idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=idx, value=None)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
        if col.dtype == "date":
            cell.number_format = "yyyy-mm-dd"

    # 코드값 컬럼 드롭다운 + 날짜 서식 (데이터 입력 구간 2~500행)
    last_row = 500
    for idx, col in enumerate(cols, start=1):
        letter = get_column_letter(idx)
        rng = f"{letter}2:{letter}{last_row}"
        if col.dtype == "code" and col.allowed_values:
            # Excel 목록 검증: 값에 콤마가 없다는 전제(코드값은 보통 단어)
            formula = '"' + ",".join(col.allowed_values) + '"'
            if len(formula) <= 255:
                dv = DataValidation(
                    type="list", formula1=formula, allow_blank=not col.required,
                    showErrorMessage=True,
                )
                dv.error = f"허용값 중에서 선택하세요: {', '.join(col.allowed_values)}"
                dv.errorTitle = "허용되지 않은 코드값"
                dv.prompt = f"{'/'.join(col.allowed_values)} 중 선택"
                dv.promptTitle = col.name
                ws.add_data_validation(dv)
                dv.add(rng)
        elif col.dtype == "date":
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=idx).number_format = "yyyy-mm-dd"
        elif col.dtype == "number":
            dv = DataValidation(
                type="custom",
                formula1=f'OR({letter}2="",ISNUMBER({letter}2))',
                allow_blank=not col.required,
                showErrorMessage=True,
            )
            dv.error = "숫자 형식으로 입력하세요."
            dv.errorTitle = "숫자 형식 오류"
            ws.add_data_validation(dv)
            dv.add(rng)

    ws.freeze_panes = "A2"

    # 작성안내 시트
    guide = wb.create_sheet("작성안내")
    guide["A1"] = spec.title
    guide["A1"].font = Font(bold=True, size=14)
    row = 2
    if spec.purpose:
        guide.cell(row=row, column=1, value="목적")
        guide.cell(row=row, column=2, value=spec.purpose)
        row += 1
    if spec.deadline:
        guide.cell(row=row, column=1, value="제출 기한")
        guide.cell(row=row, column=2, value=spec.deadline)
        row += 1
    row += 1

    headers = ["컬럼", "필수", "형식", "허용값", "예시", "설명"]
    for j, h in enumerate(headers, start=1):
        c = guide.cell(row=row, column=j, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1
    type_label = {"text": "텍스트", "date": "날짜(YYYY-MM-DD)", "number": "숫자", "code": "코드값"}
    for col in cols:
        guide.cell(row=row, column=1, value=col.name)
        guide.cell(row=row, column=2, value="필수" if col.required else "선택")
        guide.cell(row=row, column=3, value=type_label.get(col.dtype, col.dtype))
        guide.cell(row=row, column=4, value=", ".join(col.allowed_values))
        guide.cell(row=row, column=5, value=col.example or "")
        guide.cell(row=row, column=6, value=col.description or "")
        row += 1
    row += 1
    if spec.duplicate_keys:
        guide.cell(row=row, column=1, value="중복 판정 키")
        guide.cell(row=row, column=2, value=", ".join(spec.duplicate_keys))
        row += 1
    for note in spec.notes:
        guide.cell(row=row, column=1, value="주의")
        guide.cell(row=row, column=2, value=note)
        row += 1

    for j, w in enumerate([16, 8, 18, 24, 16, 40], start=1):
        guide.column_dimensions[get_column_letter(j)].width = w

    wb.save(path)
    return path


# ---------- 영속화 (요청ID별 스펙 + 엑셀 저장) ----------

def _safe_filename(title: str) -> str:
    base = re.sub(r"[^0-9A-Za-z가-힣_\- ]", "", title).strip() or "취합양식"
    return base.replace(" ", "_") + ".xlsx"


def save_template_spec(spec: TemplateSpec, template_id: str) -> Path:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    p = TEMPLATE_DIR / f"{template_id}.json"
    p.write_text(spec.model_dump_json(indent=2), encoding="utf-8")
    return p


def load_template_spec(template_id: str) -> Optional[TemplateSpec]:
    p = TEMPLATE_DIR / f"{template_id}.json"
    if not p.exists():
        return None
    try:
        return TemplateSpec.model_validate_json(p.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return None


def template_excel_path(template_id: str) -> Optional[Path]:
    """저장된 양식 엑셀 파일 경로(있으면)."""
    d = TEMPLATE_DIR / template_id
    if not d.exists():
        return None
    xlsx = sorted(d.glob("*.xlsx"))
    return xlsx[0] if xlsx else None


def build_and_save_template(spec: TemplateSpec) -> dict:
    """스펙을 저장하고 엑셀 양식을 생성한 뒤, 요약 정보를 반환한다."""
    template_id = "TPL-" + datetime.now().strftime("%Y%m%d-%H%M%S-%f")[:-3]
    save_template_spec(spec, template_id)
    filename = _safe_filename(spec.title)
    xlsx_path = write_template_excel(spec, TEMPLATE_DIR / template_id / filename)
    rule = template_spec_to_validation_rule(spec)
    return {
        "template_id": template_id,
        "filename": filename,
        "excel_path": str(xlsx_path),
        "download": f"/api/download-template/{template_id}",
        "validation_rule": rule.model_dump(),
        "template_spec": spec.model_dump(),
    }
